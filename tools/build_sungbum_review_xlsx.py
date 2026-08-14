#!/usr/bin/env python3
"""build_sungbum_review_xlsx.py — the human-adjudication workbook
for Sungbum texts the title matcher could not link: each row shows
our catalog title beside the three closest BDRC candidates with
similarity scores. Reviewers mark a decision; nothing enters the
concordance without it.

  input   data/extracted/sungbum_unmatched_review.json
  output  docs/review/Sungbum_Scan_Review.xlsx
          (+ copy to ~/Desktop/ALL Tool Distribution/)
"""
import json, os, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT,
                   "data/extracted/sungbum_unmatched_review.json")
OUTD = os.path.join(ROOT, "docs", "review")
OUT = os.path.join(OUTD, "Sungbum_Scan_Review.xlsx")
DESK = os.path.expanduser(
    "~/Desktop/ALL Tool Distribution/Sungbum_Scan_Review.xlsx")

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
HEAD = PatternFill("solid", fgColor="EEE6D8")
INPUT = PatternFill("solid", fgColor="FFFFCC")


def main():
    rows = json.load(open(SRC))
    os.makedirs(OUTD, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append(["ACIP text", "Author (BDRC person)",
               "Our catalog title (unicode)",
               "Candidate 1", "Score 1", "Nodes 1",
               "Candidate 2", "Score 2", "Nodes 2",
               "Candidate 3", "Score 3", "Nodes 3",
               "DECISION (1/2/3/none)", "Reviewer notes"])
    for c in ws[1]:
        c.font = BOLD
        c.fill = HEAD
    for r in rows:
        cands = r.get("candidates", [])
        row = [r["s"], r["pid"], r["ours"]]
        for i in range(3):
            if i < len(cands):
                row += [cands[i]["title"], cands[i]["score"],
                        cands[i]["nodes"]]
            else:
                row += ["", "", ""]
        row += ["", ""]
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = ARIAL
            c.alignment = Alignment(vertical="top",
                                    wrap_text=True)
        row[12].fill = INPUT   # decision
        row[13].fill = INPUT   # notes
    widths = [9, 10, 44, 44, 7, 24, 44, 7, 24, 44, 7, 24, 12, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    # legend sheet
    lg = wb.create_sheet("How to review")
    lines = [
        "SUNGBUM SCAN-LINK REVIEW — machine proposes, you decide",
        "",
        "Each row is an ACIP Sungbum text the title matcher could",
        "not link to a BDRC scan node with certainty. The three",
        "closest BDRC titles are shown with similarity scores",
        "(1.000 = identical after normalization).",
        "",
        "Fill only the YELLOW cells:",
        "  DECISION — 1, 2, or 3 to accept that candidate;",
        "             'none' if no candidate is the same text.",
        "  Reviewer notes — anything the next person needs.",
        "",
        "Accepted decisions are merged into the concordance with",
        "tier 'human-reviewed'; 'none' rows stay honestly",
        "unlinked. Nothing ships without a decision.",
    ]
    for i, t in enumerate(lines, 1):
        cell = lg.cell(row=i, column=1, value=t)
        cell.font = BOLD if i == 1 else ARIAL
    lg.column_dimensions["A"].width = 64
    wb.save(OUT)
    try:
        shutil.copy(OUT, DESK)
    except Exception:
        pass
    print(f"{len(rows)} rows -> {OUT} (+ Desktop copy)")


if __name__ == "__main__":
    main()
